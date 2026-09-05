from __future__ import annotations

import os
from io import BytesIO
from pathlib import Path
import tempfile
from typing import Any

import openpyxl

from model import MEMORY_TYPES, InputFormatError, MemoryShape, parse_int


SHEET_NAME = "memory_list"
EXCEL_HEADERS = (
    "mem_type",
    "prefix",
    "suffix",
    "depth",
    "width",
    "strb_w",
    "mem_user",
    "wr_clk_MHz",
    "rd_clk_MHz",
    "ppa_target",
    "instance_num",
    "capacity_KiB",
    "hierarchy",
)
REQUIRED_HEADERS = EXCEL_HEADERS[:7]


def _atomic_save_workbook(workbook: openpyxl.Workbook, path: Path) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    descriptor, temporary_name = tempfile.mkstemp(
        prefix=f".{path.stem}.",
        suffix=path.suffix,
        dir=path.parent,
    )
    os.close(descriptor)
    temporary_path = Path(temporary_name)
    try:
        workbook.save(temporary_path)
        os.replace(temporary_path, path)
    finally:
        temporary_path.unlink(missing_ok=True)


def _header_map(worksheet: openpyxl.worksheet.worksheet.Worksheet) -> dict[str, int]:
    result: dict[str, int] = {}
    for column in range(1, worksheet.max_column + 1):
        value = worksheet.cell(row=1, column=column).value
        if value is None:
            continue
        name = str(value).strip()
        if name in result:
            raise InputFormatError(f"duplicate Excel header {name!r}")
        result[name] = column
    missing = [name for name in REQUIRED_HEADERS if name not in result]
    if missing:
        raise InputFormatError(
            f"Excel sheet {SHEET_NAME!r} is missing headers: {', '.join(missing)}"
        )
    return result


def _cell_value(
    worksheet: openpyxl.worksheet.worksheet.Worksheet,
    headers: dict[str, int],
    row: int,
    name: str,
    default: Any = None,
) -> Any:
    column = headers.get(name)
    if column is None:
        return default
    value = worksheet.cell(row=row, column=column).value
    return default if value is None else value


def write_memory_excel(
    shapes_by_type: dict[str, list[MemoryShape]],
    path: Path,
    *,
    default_wr_clk_mhz: int,
    default_rd_clk_mhz: int,
) -> None:
    default_wr_clk_mhz = parse_int(
        default_wr_clk_mhz, "default_wr_clk_mhz", 1
    )
    default_rd_clk_mhz = parse_int(
        default_rd_clk_mhz, "default_rd_clk_mhz", 1
    )

    workbook = openpyxl.Workbook()
    worksheet = workbook.active
    worksheet.title = SHEET_NAME
    for column, header in enumerate(EXCEL_HEADERS, 1):
        worksheet.cell(row=1, column=column).value = header
        worksheet.column_dimensions[
            openpyxl.utils.get_column_letter(column)
        ].width = 15

    row = 2
    for mem_type in MEMORY_TYPES:
        for shape in shapes_by_type.get(mem_type, []):
            wr_clk = shape.wr_clk_mhz or default_wr_clk_mhz
            rd_clk = shape.rd_clk_mhz or (
                default_rd_clk_mhz
                if shape.mem_type == "tpram2ck"
                else wr_clk
            )
            values = {
                "mem_type": shape.mem_type,
                "prefix": shape.prefix,
                "suffix": shape.suffix,
                "depth": shape.depth,
                "width": shape.width,
                "strb_w": shape.strb_w,
                "mem_user": shape.mem_user,
                "wr_clk_MHz": wr_clk,
                "rd_clk_MHz": rd_clk,
                "ppa_target": shape.ppa_target,
                "instance_num": shape.instance_num,
                "capacity_KiB": round(shape.capacity_kib, 2),
                "hierarchy": shape.hierarchy,
            }
            for column, header in enumerate(EXCEL_HEADERS, 1):
                worksheet.cell(row=row, column=column).value = values[header]
            row += 1

    _atomic_save_workbook(workbook, path)
    workbook.close()


def parse_memory_excel(path: Path) -> dict[str, list[MemoryShape]]:
    if not path.is_file():
        raise InputFormatError(f"Excel file does not exist: {path}")
    try:
        workbook = openpyxl.load_workbook(
            BytesIO(path.read_bytes()),
            data_only=True,
        )
    except (OSError, ValueError) as exc:
        raise InputFormatError(f"cannot open Excel file {path}: {exc}") from exc
    try:
        if SHEET_NAME not in workbook.sheetnames:
            raise InputFormatError(
                f"Excel file {path} does not contain sheet {SHEET_NAME!r}"
            )
        worksheet = workbook[SHEET_NAME]
        headers = _header_map(worksheet)
        result = {mem_type: [] for mem_type in MEMORY_TYPES}
        condition_rows: dict[tuple[str, int, int, int, int], int] = {}
        for row in range(2, worksheet.max_row + 1):
            required_values = [
                _cell_value(worksheet, headers, row, name)
                for name in REQUIRED_HEADERS
            ]
            if all(value is None for value in required_values):
                continue
            values = {
                "mem_type": _cell_value(
                    worksheet, headers, row, "mem_type"
                ),
                "prefix": _cell_value(worksheet, headers, row, "prefix"),
                "suffix": _cell_value(
                    worksheet, headers, row, "suffix", ""
                ),
                "depth": _cell_value(worksheet, headers, row, "depth"),
                "width": _cell_value(worksheet, headers, row, "width"),
                "strb_w": _cell_value(worksheet, headers, row, "strb_w"),
                "mem_user": _cell_value(
                    worksheet, headers, row, "mem_user"
                ),
                "wr_clk_mhz": _cell_value(
                    worksheet, headers, row, "wr_clk_MHz"
                ),
                "rd_clk_mhz": _cell_value(
                    worksheet, headers, row, "rd_clk_MHz"
                ),
                "ppa_target": _cell_value(
                    worksheet, headers, row, "ppa_target", 0
                ),
                "instance_num": _cell_value(
                    worksheet, headers, row, "instance_num", 1
                ),
                "hierarchy": _cell_value(
                    worksheet, headers, row, "hierarchy", ""
                ),
            }
            try:
                shape = MemoryShape.from_mapping(values)
            except InputFormatError as exc:
                raise InputFormatError(f"{path}:{row}: {exc}") from exc
            condition = (
                shape.mem_type,
                shape.depth,
                shape.width,
                shape.strb_w,
                shape.mem_user,
            )
            if condition in condition_rows:
                raise InputFormatError(
                    f"{path}:{row}: duplicate memory condition; "
                    f"first defined at row {condition_rows[condition]}"
                )
            condition_rows[condition] = row
            result[shape.mem_type].append(shape)
        return result
    finally:
        workbook.close()
