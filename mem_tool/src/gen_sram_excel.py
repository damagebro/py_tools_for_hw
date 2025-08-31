import os,sys,re,json,argparse
import openpyxl
import shutil,pathlib

#0. common_func-----------------------------------------
#1. gen_sram_shell--------------------------------------
#2. gen_sram_lst_by_runsim------------------------------
#3. gen_sram_excel--------------------------------------
#4. gen_sram_instance-----------------------------------

# sram_rpt_json = {
#     "spram_lst" : [
#         { "mem_type":"spram","prefix":"","suffix":"", "raw_shape":"", "depth":1,"width":1,"strb_w":1,"mem_user":0,  "hierachey":"h1,h2", "instance_num":2 },
#         { "mem_type":"spram","prefix":"","suffix":"", "raw_shape":"", "depth":10,"width":6,"strb_w":1,"mem_user":0,  "hierachey":"h1,h2", "instance_num":2 },
#         { "mem_type":"spram","prefix":"","suffix":"", "raw_shape":"", "depth":10,"width":6,"strb_w":2,"mem_user":0,  "hierachey":"h1,h2", "instance_num":2 },
#         { "mem_type":"spram","prefix":"","suffix":"", "raw_shape":"", "depth":10,"width":6,"strb_w":1,"mem_user":1,  "hierachey":"h1,h2", "instance_num":2 },
#         { "mem_type":"spram","prefix":"","suffix":"", "raw_shape":"", "depth":10,"width":6,"strb_w":2,"mem_user":2,  "hierachey":"h1,h2", "instance_num":2 },
#     ],
#     "tpram1ck_lst" : [
#         { "mem_type":"tpram1ck","prefix":"","suffix":"", "raw_shape":"", "depth":10,"width":6,"strb_w":2,"mem_user":2,  "hierachey":"h1,h2", "instance_num":2 },
#     ],
#     "tpram2ck_lst" : [],
#     "sprom_lst" : []
# }
g_excel_col_idx = { "mem_type":1,"prefix":2,"suffix":3,  "depth":4,"width":5,"strb_w":6,"mem_user":7, "wr_clk_MHz":8,"rd_clk_MHz":9,"ppa_target":10,  "instance_num":11,"capacity_KiB":12,"hierachy":13 }
g_li_sram_type = ['spram','tpram1ck','tpram2ck','sprom']
# g_li_sram_type = ['spram','tpram1ck','tpram2ck']

#0. common_func-----------------------------------------
def str_re_find( str_ori, str_pat ):
    'ret: if str_ori have str_pat, return a integer number, which is larger than 0'
    ui_ret = len(re.findall(str_pat,str_ori,flags=0))
    return( ui_ret )
def file2lst( fn ):
    li_ret = []
    if( not os.path.exists(fn) ):
        print( f'NOTICE(), no such file or directory:"{fn}"' )
        exit(-1)
    fp = open(fn,'rt')
    for s in fp.readlines():
        li_ret.append( s.rstrip() )
    fp.close()
    #li_ret = pathlib.Path(fn).read_text(encoding='utf-8').splitlines()
    return li_ret
def list2str( li_str ):
    str_ret = ""
    for i,s in enumerate(li_str):
        s = li_str[i]
        str_ret += f'{s}\n'
    return str_ret
def li_str_split( li_str, str_pat ):
    'ret: li_prev=[0:pattern], li_match=[pattern], li_last[pattern+1:-1]'
    li_prev  = []
    li_match = []
    li_last  = []
    ui_match_idx = -1
    for i,s in enumerate(li_str):
        if( str_re_find(s, str_pat) ):
            ui_match_idx = i
            break
    li_prev.extend(li_str[0:ui_match_idx])
    if( ui_match_idx>0 ):
        li_match.extend(li_str[ui_match_idx:ui_match_idx+1])
    if( ui_match_idx>0 and ui_match_idx<len(li_str) ):
        li_last.extend(li_str[ui_match_idx+1:])
    # print( 'match_idx:{}, ori_str:\n{}\n  prev:\n{}\n  match:\n{}\n  last:\n{}\n'.format(ui_match_idx, li_str, li_prev,li_match,li_last) )
    return li_prev,li_match,li_last
def file_str_split(fn, str_pat):
    'ret: str_prev, str_match, str_last'
    li_file = file2lst(fn)
    li_prev,li_match,li_last = li_str_split(li_file, str_pat)
    str_prev  = list2str(li_prev )
    str_match = list2str(li_match)
    str_last  = list2str(li_last )
    return str_prev,str_match,str_last
def file_str_sub(fn, str_pat, str_rep):
    fp = pathlib.Path(fn)
    str_file = fp.read_text(encoding='utf-8')
    str_ret = re.sub(str_pat, str_rep, str_file, count=0, flags=0)  #count可选，最大替换次数（默认 0 表示替换所有匹配）；   flags：可选，正则匹配标志（如 re.IGNORECASE 忽略大小写）
    fp.write_text(str_ret, encoding='utf-8')

#1. gen_sram_shell--------------------------------------
def gen_sram_shell( sram_cfg_json ):
    str_work_path = sram_cfg_json['common_opt']['env_var']['work_path']
    str_ancr_path = sram_cfg_json['common_opt']['env_var']['anchor_path']
    str_shel_path = str_ancr_path + '/shell_template/'
    str_ori_prefix = 'com'
    str_new_prefix = sram_cfg_json['common_opt']['subsys_prefix']
    for sram_type in g_li_sram_type:
        ori_ram_shell_name = f'{str_ori_prefix}_{sram_type}_shell'
        new_ram_shell_name = f'{str_new_prefix}_{sram_type}_shell'
        ori_ram_file = f'{str_shel_path}/{ori_ram_shell_name}.sv'
        new_ram_file = f'{str_work_path}/{new_ram_shell_name}.sv'
        tgt_dir  = os.path.dirname(new_ram_file)
        os.makedirs(tgt_dir, exist_ok=True)  #递归创建，exist_ok=True 避免目录已存在时报错
        shutil.copy2(ori_ram_file, new_ram_file)
        file_str_sub( new_ram_file, ori_ram_shell_name,new_ram_shell_name)

        if( not str_re_find(sram_type,'sprom') ):
            ori_ecc_shell_name = f'{str_ori_prefix}_ecc_{sram_type}_shell'
            new_ecc_shell_name = f'{str_new_prefix}_ecc_{sram_type}_shell'
            ori_ecc_file = f'{str_shel_path}/{ori_ecc_shell_name}.sv'
            new_ecc_file = f'{str_work_path}/{new_ecc_shell_name}.sv'
            shutil.copy2(ori_ecc_file, new_ecc_file)
            file_str_sub( new_ecc_file, ori_ecc_shell_name,new_ecc_shell_name)
            file_str_sub( new_ecc_file, ori_ram_shell_name,new_ram_shell_name)
        if( str_re_find(sram_type,'sprom') ):
            ori_name = f'{str_ori_prefix}_{sram_type}_mannul'
            new_name = f'{str_new_prefix}_{sram_type}_mannul'
            ori_file = f'{str_shel_path}/{ori_name}.sv'
            new_file = f'{str_work_path}/{new_name}.sv'
            shutil.copy2(ori_file, new_file)
            file_str_sub( new_file, ori_name,new_name)
            file_str_sub( new_file, ori_ram_shell_name,new_ram_shell_name)
    print( f'{str_new_prefix}_sram_shell without sram_shape intergrated, generate successfully in the path "{str_work_path}"' )

#2. gen_sram_lst_by_runsim------------------------------
def run_vcs_sim(sram_cfg_json):
    pass
#3. gen_sram_excel--------------------------------------
def parse_one_sram_lst(sram_cfg_json, fn):
    'ret: li_ram_lst'
    # {"spram_lst" : li_ram_lst=[ { "raw_shape":"", "depth":1,"width":1,"strb_w":1,"mem_user":0, ..},... ] "tpram1ck_lst" : li_ram_lst, "tpram2ck_lst" : li_ram_lst, "sprom_lst" : li_ram_lst}
    # g_excel_col_idx = { "mem_type":1,"prefix":2,"suffix":3,  "depth":4,"width":5,"strb_w":6,"mem_user":7, "wr_clk/MHz":8,"rd_clk/MHz":9,"ppa_target":10,  "instance_num":11,"capacity/KiB":12,"hierachy":13 }
    li_ram = []
    li_ram_unique = []
    li_ram_lst = []
    str_prefix = sram_cfg_json['common_opt']['subsys_prefix']
    for s in file2lst(fn):  # li_file = pathlib.Path(fn).read_text(encoding='utf-8').splitlines()
        li_tmp = s.split()  #[0=shape, 1=warning/info/message, -1=hierachy]
        if( len(li_tmp)<=1 ):
            continue
        dict_tmp = {}
        str_shape = li_tmp[0]
        str_hie   = li_tmp[-1]
        ram_type = ''
        for sram_type in g_li_sram_type:
            if( str_re_find(str_shape, sram_type) ):
                ram_type = sram_type
                break
        dict_tmp['raw_shape'] = str_shape
        dict_tmp['mem_type']  = ram_type
        dict_tmp['prefix']    = str_prefix
        ui_user = 0
        if( str_re_find(str_shape, '_usr') ):
            ui_user = int(str_shape.split(ram_type)[-1].split('_usr')[-1])
            dict_tmp['suffix'] = f'usr{ui_user}'
        else:
            dict_tmp['suffix'] = ""
        str_shape_tmp = str_shape.split(ram_type)[-1].split('_usr')[0]   #depthxwidth[xstrb_w]
        dict_tmp['depth'] = int(str_shape_tmp.split('x')[0])
        dict_tmp['width'] = int(str_shape_tmp.split('x')[1])
        dict_tmp['strb_w']= 1
        if( len(str_shape_tmp.split('x'))>2 ):
            dict_tmp['strb_w'] = int(str_shape_tmp.split('x')[2])
        dict_tmp['mem_user'] = ui_user
        dict_tmp['hierachy'] = str_hie
        li_ram.append(dict_tmp)
    #ram_shape unique---------------
    for d1 in li_ram:
        str_raw_shape = d1['raw_shape']
        b_new_flag = 1
        for d2 in li_ram_unique:
            exist_shape = d2['raw_shape']
            if( str_raw_shape==exist_shape ):
                b_new_flag = 0
                break
        if( b_new_flag ):
            li_ram_unique.append(d1)
    #instance_num---------------
    for d1 in li_ram_unique:
        unique_shape = d1['raw_shape']
        ui_cnt = 0
        str_hie_all = ''
        for d2 in li_ram:
            verbose_shape = d2['raw_shape']
            if( verbose_shape==unique_shape ):
                ui_cnt +=1
                if( str_hie_all=='' ):
                    str_hie_all = d2['hierachy']
                else:
                    str_hie_all = f'{str_hie_all},{d2['hierachy']}'
        d1['instance_num'] = ui_cnt
        ui_tol_capacity_bit = d1['depth']*d1['width']*ui_cnt
        d1['capacity_KiB'] = '{:.2f}'.format(ui_tol_capacity_bit/8/1024)
        d1['hierachy'] = str_hie_all
        li_ram_lst.append(d1)
    return li_ram_lst
def parse_all_sram_lst(sram_cfg_json):
    'ret: sram_rpt_json = {"spram_lst":[] "tpram1ck_lst":[], "tpram2ck_lst":[], "sprom_lst":[]}'
    str_rpt_path = sram_cfg_json['common_opt']['env_var']['work_path']
    sram_rpt_json = {}
    for sram_type in g_li_sram_type:
        li_ram_lst = []
        rpt_filename = f'{str_rpt_path}/{sram_type}.lst'
        if( os.path.exists(rpt_filename) ):
            li_ram_lst = parse_one_sram_lst( sram_cfg_json, rpt_filename )
        sram_rpt_json[f'{sram_type}_lst'] = li_ram_lst
    #把 mem_type 放到dict[mem_type_lst]中
    sram_rpt_json_ret = {}
    for sram_type in g_li_sram_type:
        sram_rpt_json_ret[f'{sram_type}_lst'] = []
    for sram_type in g_li_sram_type:
        li_ram_lst = sram_rpt_json[f'{sram_type}_lst']
        for dict_item in li_ram_lst:
            mem_type = dict_item['mem_type']
            sram_rpt_json_ret[f'{mem_type}_lst'].append(dict_item)
    return sram_rpt_json_ret

def gen_sram_excel( sram_cfg_json ):
    # g_excel_col_idx = { "mem_type":1,"prefix":2,"suffix":3,  "depth":4,"width":5,"strb_w":6,"mem_user":7, "wr_clk_MHz":8,"rd_clk_MHz":9,"ppa_target":10,  "instance_num":11,"capacity_KiB":12,"hierachy":13 }
    str_work_path = sram_cfg_json['common_opt']['env_var']['work_path']
    wb_name = str_work_path + sram_cfg_json['common_opt']['excel_filename']
    ws_name = "memory_list"
    tgt_dir  = os.path.dirname(wb_name)
    os.makedirs(tgt_dir, exist_ok=True)

    #1. row[1], each title;
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = ws_name
    for k,v in g_excel_col_idx.items():
        col_letter = openpyxl.utils.get_column_letter(v)
        ws.column_dimensions[col_letter].width = 15
        ws.cell(row=1, column=v).value = k
    # wb.save(wb_name)

    #2. get all sram_lst from sram_rpt_json;
    sram_rpt_json = parse_all_sram_lst(sram_cfg_json); #print( json.dumps(sram_rpt_json, indent=4) )
    for sram_type in g_li_sram_type:
        li_ram_lst = sram_rpt_json[f'{sram_type}_lst']
        ui_row_idx = ws.max_row
        for dict_item in li_ram_lst:
            ui_row_idx += 1
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['mem_type']).value = dict_item['mem_type']
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['prefix'  ]).value = dict_item['prefix'  ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['suffix'  ]).value = dict_item['suffix'  ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['depth'   ]).value = dict_item['depth'   ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['width'   ]).value = dict_item['width'   ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['strb_w'  ]).value = dict_item['strb_w'  ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['mem_user']).value = dict_item['mem_user']
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['instance_num']).value = dict_item['instance_num']
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['capacity_KiB']).value = dict_item['capacity_KiB']
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['hierachy'    ]).value = dict_item['hierachy'    ]
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['wr_clk_MHz'  ]).value = sram_cfg_json['gen_sram_excel']['default_ram_wr_clk_MHz']
            ws.cell(row=ui_row_idx, column=g_excel_col_idx['ppa_target'  ]).value = 0   #TBD
            if( dict_item['mem_user']>0 ):
                # ws.cell(ui_row_idx=ui_row_idx, column=g_excel_col_idx['wr_clk_MHz']).value = func_user2clk(dict_item['mem_user'])  #TBD
                pass
            if( dict_item['mem_type']=='tpram2ck' ):
                ws.cell(row=ui_row_idx, column=g_excel_col_idx['rd_clk_MHz']).value = sram_cfg_json['gen_sram_excel']['default_ram_rd_clk_MHz']
            else:
                ws.cell(row=ui_row_idx, column=g_excel_col_idx['rd_clk_MHz']).value = ws.cell(row=ui_row_idx, column=g_excel_col_idx['wr_clk_MHz']).value
    wb.save(wb_name)
    print( f'sram_excel generate successfully in the path "{wb_name}"' )

#4. gen_sram_instance-----------------------------------
def parse_sram_excel( fn ):
    wb = openpyxl.load_workbook(fn)
    ws = wb.active
    sram_rpt_json = {}  #{"spram_lst":[] "tpram1ck_lst":[], "tpram2ck_lst":[], "sprom_lst":[]}
    for row_idx in range(2,ws.max_row):
        dict_tmp = {}
        dict_tmp['mem_type'] = ws.cell(row=row_idx, column=g_excel_col_idx['mem_type']).value
        dict_tmp['prefix'  ] = ws.cell(row=row_idx, column=g_excel_col_idx['prefix'  ]).value
        dict_tmp['suffix'  ] = ws.cell(row=row_idx, column=g_excel_col_idx['suffix'  ]).value
        dict_tmp['depth'   ] = ws.cell(row=row_idx, column=g_excel_col_idx['depth'   ]).value
        dict_tmp['width'   ] = ws.cell(row=row_idx, column=g_excel_col_idx['width'   ]).value
        dict_tmp['strb_w'  ] = ws.cell(row=row_idx, column=g_excel_col_idx['strb_w'  ]).value
        dict_tmp['mem_user'] = ws.cell(row=row_idx, column=g_excel_col_idx['mem_user']).value
        if f'{dict_tmp['mem_type']}_lst' not in sram_rpt_json:
            sram_rpt_json[f'{dict_tmp['mem_type']}_lst'] = []
        sram_rpt_json[f'{dict_tmp['mem_type']}_lst'].append(dict_tmp)
    return sram_rpt_json
def gen_sram_instance( sram_cfg_json ):
    str_work_path = sram_cfg_json['common_opt']['env_var']['work_path']
    str_ancr_path = sram_cfg_json['common_opt']['env_var']['anchor_path']
    str_shel_path = str_ancr_path + '/shell_template/'
    str_ori_prefix = 'com'
    str_new_prefix = sram_cfg_json['common_opt']['subsys_prefix']

    dict_inst = {}
    dict_inst['spram']    = "( .clk(clk),.sys_cfg(sys_cfg), .wr_en(wr_en),.wr_addr(wr_addr),.wr_data(wr_data), .rd_en(rd_en),.rd_addr(rd_addr),.rd_data(rd_data) );"
    dict_inst['tpram1ck'] = "( .clk(clk),.sys_cfg(sys_cfg), .wr_en(wr_en),.wr_addr(wr_addr),.wr_data(wr_data), .rd_en(rd_en),.rd_addr(rd_addr),.rd_data(rd_data) );"
    dict_inst['tpram2ck'] = "( .clk(clk),.sys_cfg(sys_cfg), .wr_en(wr_en),.wr_addr(wr_addr),.wr_data(wr_data), .rd_en(rd_en),.rd_addr(rd_addr),.rd_data(rd_data) );"
    dict_inst['sprom']    = "( .clk(clk),.sys_cfg(sys_cfg), .rd_en(rd_en),.rd_addr(rd_addr),.rd_data(rd_data) );"

    wb_name = str_work_path + sram_cfg_json['common_opt']['excel_filename']
    if( os.path.isfile(wb_name) ):
        sram_rpt_json = parse_sram_excel(wb_name)
    else:
        sram_rpt_json = parse_all_sram_lst(sram_cfg_json)
    #print( json.dumps(sram_rpt_json, indent=4) )
    for k,v in sram_rpt_json.items():
        sram_type = k.split('_lst')[0]
        li_ram = v
        #1. copy sram_shell+ecc_shell
        ori_ram_shell_name = f'{str_ori_prefix}_{sram_type}_shell'
        new_ram_shell_name = f'{str_new_prefix}_{sram_type}_shell'
        ori_ram_file = f'{str_shel_path}/{ori_ram_shell_name}.sv'
        new_ram_file = f'{str_work_path}/{new_ram_shell_name}.sv'
        tgt_dir  = os.path.dirname(new_ram_file)
        os.makedirs(tgt_dir, exist_ok=True)  #递归创建，exist_ok=True 避免目录已存在时报错
        shutil.copy2(ori_ram_file, new_ram_file)
        file_str_sub( new_ram_file, ori_ram_shell_name,new_ram_shell_name)

        if( not str_re_find(sram_type,'sprom') ):
            ori_ecc_shell_name = f'{str_ori_prefix}_ecc_{sram_type}_shell'
            new_ecc_shell_name = f'{str_new_prefix}_ecc_{sram_type}_shell'
            ori_ecc_file = f'{str_shel_path}/{ori_ecc_shell_name}.sv'
            new_ecc_file = f'{str_work_path}/{new_ecc_shell_name}.sv'
            shutil.copy2(ori_ecc_file, new_ecc_file)
            file_str_sub( new_ecc_file, ori_ecc_shell_name,new_ecc_shell_name)
            file_str_sub( new_ecc_file, ori_ram_shell_name,new_ram_shell_name)
        if( str_re_find(sram_type,'sprom') ):
            ori_name = f'{str_ori_prefix}_{sram_type}_mannul'
            new_name = f'{str_new_prefix}_{sram_type}_mannul'
            ori_file = f'{str_shel_path}/{ori_name}.sv'
            new_file = f'{str_work_path}/{new_name}.sv'
            shutil.copy2(ori_file, new_file)
            file_str_sub( new_file, ori_name,new_name)
            file_str_sub( new_file, ori_ram_shell_name,new_ram_shell_name)
        #2. gen sram_instance
        str_inst = ''
        for i,ram_item in enumerate(li_ram):
            s = f'if( DEPTH=={ram_item['depth']} && DATA_W=={ram_item['width']} && STRB_W=={ram_item['strb_w']} && MEM_USER=={ram_item['mem_user']} )begin:gen_sram_phy\n'
            if( i>0 ):
                s = f'else {s}'
            str_shape = f'{ram_item['depth']}x{ram_item['width']}'
            if( ram_item['strb_w']>1 ):
                str_shape = f'{str_shape}x{ram_item['strb_w']}'
            str_suffix = ''
            if( len(ram_item['suffix'])>1 ):
                str_suffix = f'_{str_suffix}'
            str_wrap_name = f'{ram_item['prefix']}_{sram_type}_{str_shape}{str_suffix}_wrapper'
            str_inst += f"{' '*8 }" + s
            str_inst += f"{' '*12}" + f'{str_wrap_name} u_{str_wrap_name} {dict_inst[sram_type]}\n'
            str_inst += f"{' '*12}" + "assign use_cell = 1'b1;\n"
            str_inst += f"{' '*8 }" + "end:gen_sram_phy\n"
        s_prev,s_midd,s_last = file_str_split( new_ram_file, "// Start of user logic." )
        e_prev,e_midd,e_last = file_str_split( new_ram_file, "// End of user logic." )
        str_wr = f'{s_prev}{s_midd}{str_inst}{e_midd}{e_last}'
        pathlib.Path(new_ram_file).write_text(str_wr)
    print( f'{str_new_prefix}_sram_shell with all sram_shape intergrated, generate successfully in the path "{str_work_path}"' )

#main + cfg_parser------------
def cfg_parser():
    'ret sram_cfg_json'
    sram_cfg_json = {
        "common_opt" : {
            "gen_sram_shell"    : 0,
            "gen_sram_list"     : 0,
            "gen_sram_excel"    : 0,
            "gen_sram_instance" : 0,

            "subsys_prefix" : "cpu",
            "excel_filename" : "project_cpu_20250730_memory_require.xlsx",
            "env_var" : {
                "anchor_path" : "./",
                "work_path"   : "./bin/",
                "PROJ_ROOT"   : ""
            },
        },
        "gen_sram_excel" : {
            "default_ram_wr_clk_MHz" : 1500,
            "default_ram_rd_clk_MHz" : 1000
        },
    }

    parser = argparse.ArgumentParser(
        epilog=''' 使用示例:
               (1)只生成sram_shell: python gen_sram_excel.py -p cpu [-w ./out/];   #shell输出路径是: ./out/
               (2)跑仿真得到sram_rpt: TBD;
               (3)生成sram_excel: python gen_sram_excel.py -p cpu -m excel -w ./bin/ -x cpu_20250730_memory_require.xlsx -xcka 1500 [-xckb 1000]
                   #rpt输入路径是: ./bin/spram.lst,./bin/tpram1ck.lst...;
                   #excel输出路径是: ./bin/cpu_20250730_memory_require.xlsx;
               (4)例化instance: python gen_sram_excel.py -p cpu -m inst -w ./bin/ {-x cpu_20250730_memory_require.xlsx|}
                   #shell输出路径是: ./bin/;
                   #如果有-x,则从excel中得到所有sram_list然后instance;
                   #如果没有-x, 则./bin/spram.lst,./bin/tpram1ck.lst... 得到所有sram_list然后instance;
               '''
    )
    parser.add_argument('-p', '--subsys_preifx', required=True, help='subsys_prefix name,  [cpu|npu|pcie...]')
    parser.add_argument('-m', '--mode', default='init', help='work_mode = {init|inst|excel|rpt_by_run_sim}')
    parser.add_argument('-w', '--work_path', default='./')
    parser.add_argument('-x', '--excel_name', help='sram_excel_name, like cpu_20250730_memory_require.xlsx')
    parser.add_argument('-xcka', '--clk_a', help='clka_freq_MHz, clk_a=spram/tpram1ck/tpram2ck.wr_clk')
    parser.add_argument('-xckb', '--clk_b', help='clkb_freq_MHz, clk_b=tprambck.rd_clk')
    args = parser.parse_args()
    sram_cfg_json['common_opt']['subsys_prefix'] = args.subsys_preifx
    if( args.mode=='init' ):
        sram_cfg_json['common_opt']['gen_sram_shell'] = 1
    if( args.mode=='inst' ):
        sram_cfg_json['common_opt']['gen_sram_instance'] = 1
        if( args.excel_name is not None ):
            sram_cfg_json['common_opt']['excel_filename'] = args.excel_name
        else:
            sram_cfg_json['common_opt']['excel_filename'] = ''
    if( args.mode=='excel' ):
        sram_cfg_json['common_opt']['gen_sram_excel'] = 1
        sram_cfg_json['common_opt']['excel_filename'] = args.excel_name
        sram_cfg_json['gen_sram_excel']['default_ram_wr_clk_MHz'] = args.clk_a
        sram_cfg_json['gen_sram_excel']['default_ram_rd_clk_MHz'] = args.clk_a
        if( args.clk_b  is not None ):
            sram_cfg_json['gen_sram_excel']['default_ram_rd_clk_MHz'] = args.clk_b
    if( args.mode=='rpt_by_run_sim' ):
        sram_cfg_json['common_opt']['gen_sram_list'] = 1

    sram_cfg_json['common_opt']['env_var']['work_path']   = args.work_path
    sram_cfg_json['common_opt']['env_var']['anchor_path'] = './'
    # sram_cfg_json['common_opt']['env_var']['PROJ_ROOT']   = ''

    print( 'argv parse successfully' )
    return sram_cfg_json


if(1):
    # fn = 'sram_cfg.json'
    # fp = pathlib.Path(fn)
    # sram_cfg_json = json.loads(fp.read_text(encoding='utf-8'))
    sram_cfg_json = cfg_parser()
    if( sram_cfg_json['common_opt']['gen_sram_shell'] and not sram_cfg_json['common_opt']['gen_sram_instance'] ):
        gen_sram_shell(sram_cfg_json)
    if( sram_cfg_json['common_opt']['gen_sram_list'] ):
        run_vcs_sim()
    if( sram_cfg_json['common_opt']['gen_sram_excel'] ):
        gen_sram_excel(sram_cfg_json)
    if( sram_cfg_json['common_opt']['gen_sram_instance'] ):
        gen_sram_instance(sram_cfg_json)
