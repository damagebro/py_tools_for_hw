from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Any

from .models import (
    BaseInfoModel,
    FieldModel,
    ModuleModel,
    RegisterModel,
    SpecialOptions,
    SubModuleNode,
)
from .reg_common import (
    CSRValidationError,
    expand_defaults,
    parse_int,
    parse_optional_int,
    parse_special,
    validate_identifier,
)


REQUIRED_COLUMNS = {
    "offset",
    "reg_name",
    "field",
    "msb",
    "lsb",
    "sw_access",
    "default_value",
    "reg_type",
    "special",
    "description",
}

TYPE_ACCESS = {
    "cfg": "RW",
    "status": "RO",
    "cmd": "W1T",
    "toggle": "W1T",
    "irq": "W1C",
    "slave": "",
    "mem": "",
}

BASE_ALIASES = {
    "system_addr": "system_baseaddr",
    "system_base_addr": "system_baseaddr",
    "system_size": "system_bytesize",
    "system_byte_size": "system_bytesize",
}


class CSRParser:
    def __init__(self, input_path: str, nested: bool = False):
        self.input_path = Path(input_path).resolve()
        self.nested = nested
        self._active_paths: list[Path] = []

    def parse(self) -> ModuleModel:
        return self._parse_file(self.input_path, None)

    def _parse_file(self, path: Path, allocated_size: int | None) -> ModuleModel:
        path = path.resolve()
        if not path.exists():
            raise FileNotFoundError(f"Input file not found: {path}")
        if path in self._active_paths:
            chain = " -> ".join(item.name for item in self._active_paths + [path])
            raise CSRValidationError(f"Recursive slave reference detected: {chain}")

        self._active_paths.append(path)
        try:
            base_data, register_rows = self._load_input(path)
            module = ModuleModel(
                name=validate_identifier(
                    ModuleModel.clean_name(str(path)),
                    f"{path.name} module name",
                    include_c=False,
                ),
                source_path=str(path),
                base_info=self._parse_base_info(base_data, path),
            )
            self._parse_registers(module, register_rows)
            limit = allocated_size
            if module.base_info.system_bytesize is not None:
                limit = min(
                    item for item in (limit, module.base_info.system_bytesize)
                    if item is not None
                )
            if limit is not None and module.local_size > limit:
                last = module.registers[-1]
                raise CSRValidationError(
                    f"{path.name}: address space 0x{module.local_size:X} exceeds "
                    f"allocated size 0x{limit:X}; last register "
                    f"'{last.raw_name}' at offset 0x{last.offset:X}"
                )
            if self.nested:
                self._load_children(module, path)
            return module
        finally:
            self._active_paths.pop()

    def _load_input(self, path: Path) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        suffix = path.suffix.lower()
        if suffix == ".md":
            return self._load_markdown(path)
        if suffix == ".xlsx":
            return self._load_excel(path)
        raise CSRValidationError(
            f"{path.name}: unsupported input format '{path.suffix}'"
        )

    def _load_markdown(
        self, path: Path
    ) -> tuple[dict[str, str], list[dict[str, str]]]:
        sections: dict[str, list[str]] = {}
        current = ""
        for line in path.read_text(encoding="utf-8-sig").splitlines():
            stripped = line.strip()
            if stripped.startswith("#"):
                current = stripped.lstrip("#").strip().lower()
                sections.setdefault(current, [])
            elif current:
                sections[current].append(line)

        if "reg_define" not in sections:
            raise CSRValidationError(f"{path.name}: missing '# reg_define' section")
        base_headers, base_rows = self._parse_md_table(
            sections.get("base_info", []),
            path,
            "base_info",
            required=False,
        )
        reg_headers, reg_rows = self._parse_md_table(
            sections["reg_define"],
            path,
            "reg_define",
            required=True,
        )
        base_data = {
            row[base_headers[0]]: row[base_headers[1]]
            for row in base_rows
            if row.get(base_headers[0], "")
        } if len(base_headers) >= 2 else {}
        return base_data, [
            {key.lower(): value for key, value in row.items()}
            for row in reg_rows
        ]

    def _parse_md_table(
        self,
        lines: list[str],
        path: Path,
        section: str,
        required: bool,
    ) -> tuple[list[str], list[dict[str, str]]]:
        raw_rows: list[list[str]] = []
        for line in lines:
            stripped = line.strip()
            if not stripped.startswith("|"):
                continue
            cells = [item.strip() for item in stripped.strip("|").split("|")]
            if all(set(item) <= {"-", ":", " "} for item in cells):
                continue
            raw_rows.append(cells)
        if not raw_rows:
            if required:
                raise CSRValidationError(
                    f"{path.name}: section '{section}' has no table"
                )
            return [], []
        headers = raw_rows[0]
        rows = []
        for row_index, cells in enumerate(raw_rows[1:], start=1):
            cells += [""] * (len(headers) - len(cells))
            rows.append({
                **dict(zip(headers, cells[: len(headers)])),
                "__row__": str(row_index),
            })
        return headers, rows

    def _load_excel(
        self, path: Path
    ) -> tuple[dict[str, Any], list[dict[str, Any]]]:
        try:
            import openpyxl
        except ImportError as exc:
            raise ImportError(
                "openpyxl is required for .xlsx input"
            ) from exc
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            if "reg_define" not in workbook.sheetnames:
                raise CSRValidationError(
                    f"{path.name}: workbook requires a 'reg_define' sheet"
                )
            base_data: dict[str, Any] = {}
            if "base_info" in workbook.sheetnames:
                rows = self._worksheet_values(workbook["base_info"])
                for row in rows[1:]:
                    if row and row[0] not in (None, ""):
                        base_data[str(row[0]).strip()] = (
                            row[1] if len(row) > 1 else ""
                        )
            reg_values = self._worksheet_values(workbook["reg_define"])
            if not reg_values:
                raise CSRValidationError(
                    f"{path.name}: 'reg_define' sheet is empty"
                )
            headers = [
                str(item or "").strip().lower() for item in reg_values[0]
            ]
            register_rows = []
            for row_index, values in enumerate(reg_values[1:], start=1):
                if not any(item not in (None, "") for item in values):
                    continue
                padded = list(values) + [""] * (len(headers) - len(values))
                register_rows.append({
                    **dict(zip(headers, padded[: len(headers)])),
                    "__row__": row_index,
                })
            return base_data, register_rows
        finally:
            workbook.close()

    @staticmethod
    def _worksheet_values(sheet: object) -> list[tuple[Any, ...]]:
        return [
            tuple(cell.value for cell in row)
            for row in sheet.iter_rows()
        ]

    def _parse_base_info(
        self, data: dict[str, Any], path: Path
    ) -> BaseInfoModel:
        normalized = {
            BASE_ALIASES.get(str(key).strip().lower(), str(key).strip().lower()): value
            for key, value in data.items()
        }
        known = {
            "reg_bitwidth",
            "system_baseaddr",
            "system_bytesize",
            "system_prefix",
            "author",
            "email",
        }
        bitwidth = parse_int(
            normalized.get("reg_bitwidth", 32),
            f"{path.name} reg_bitwidth",
        )
        if bitwidth < 8 or bitwidth > 64 or bitwidth % 8:
            raise CSRValidationError(
                f"{path.name}: reg_bitwidth must be 8..64 and byte aligned"
            )
        return BaseInfoModel(
            reg_bitwidth=bitwidth,
            system_baseaddr=parse_optional_int(
                normalized.get("system_baseaddr", 0),
                f"{path.name} system_baseaddr",
            ) or 0,
            system_bytesize=parse_optional_int(
                normalized.get("system_bytesize"),
                f"{path.name} system_bytesize",
            ),
            system_prefix=str(normalized.get("system_prefix", "")).strip().lower(),
            author=str(normalized.get("author", "")).strip(),
            email=str(normalized.get("email", "")).strip(),
            extras={
                key: str(value)
                for key, value in normalized.items()
                if key not in known
            },
        )

    def _parse_registers(
        self, module: ModuleModel, rows: list[dict[str, Any]]
    ) -> None:
        if not rows:
            raise CSRValidationError(f"{Path(module.source_path).name}: no registers")
        columns = {str(key).lower() for key in rows[0] if not str(key).startswith("__")}
        missing = REQUIRED_COLUMNS - columns
        if missing:
            raise CSRValidationError(
                f"{Path(module.source_path).name}: missing columns: "
                + ", ".join(sorted(missing))
            )

        raw_registers: list[tuple[dict[str, Any], list[dict[str, Any]]]] = []
        current_header: dict[str, Any] | None = None
        current_fields: list[dict[str, Any]] = []
        for row in rows:
            offset = self._cell(row, "offset")
            name = self._cell(row, "reg_name")
            if offset or name:
                if current_header is not None:
                    raw_registers.append((current_header, current_fields))
                current_header = row
                current_fields = []
            elif current_header is None:
                raise CSRValidationError(
                    f"{Path(module.source_path).name} row {row.get('__row__')}: "
                    "field row appears before a register"
                )
            if self._cell(row, "field"):
                current_fields.append(row)
        if current_header is not None:
            raw_registers.append((current_header, current_fields))

        name_totals = Counter(
            self._cell(header, "reg_name").lower()
            for header, _ in raw_registers
        )
        name_indexes: defaultdict[str, int] = defaultdict(int)
        next_offset = 0
        previous: RegisterModel | None = None

        for header, field_rows in raw_registers:
            row_number = int(header.get("__row__", 0))
            raw_name = self._cell(header, "reg_name")
            if not raw_name:
                raise CSRValidationError(
                    f"{Path(module.source_path).name} row {row_number}: "
                    "reg_name is required for a new register"
                )
            base_name = validate_identifier(
                raw_name,
                f"{Path(module.source_path).name} row {row_number} reg_name",
            )
            name_indexes[base_name] += 1
            unique_name = (
                f"{base_name}{name_indexes[base_name]}"
                if name_totals[base_name] > 1 else base_name
            )
            offset_text = self._cell(header, "offset")
            offset = (
                parse_int(offset_text, f"{raw_name} offset")
                if offset_text else next_offset
            )
            if offset % module.word_bytes:
                raise CSRValidationError(
                    f"{raw_name}: offset 0x{offset:X} is not aligned to "
                    f"{module.word_bytes} bytes"
                )
            if previous is not None and offset < next_offset:
                raise CSRValidationError(
                    f"Address overlap: '{raw_name}' at 0x{offset:X}; previous "
                    f"'{previous.raw_name}' at 0x{previous.offset:X} ends at "
                    f"0x{next_offset - 1:X}"
                )

            reg_type = self._cell(header, "reg_type").lower()
            if reg_type == "toggle":
                reg_type = "cmd"
            if reg_type not in TYPE_ACCESS:
                raise CSRValidationError(
                    f"{raw_name}: reg_type '{reg_type}' is invalid"
                )
            access = self._cell(header, "sw_access").upper()
            expected_access = TYPE_ACCESS[reg_type]
            if access and access != expected_access:
                raise CSRValidationError(
                    f"{raw_name}: SW_access must be {expected_access or 'empty'} "
                    f"for reg_type={reg_type}"
                )
            access = expected_access
            special = parse_special(
                self._cell(header, "special"),
                f"{raw_name} special",
            )
            self._validate_special(raw_name, reg_type, special)

            reg = RegisterModel(
                name=unique_name,
                raw_name=raw_name,
                offset=offset,
                reg_type=reg_type,
                sw_access=access,
                special=special,
                description=self._cell(header, "description"),
                source_row=row_number,
            )
            if reg_type not in {"slave", "mem"}:
                if not field_rows:
                    raise CSRValidationError(f"{raw_name}: at least one field is required")
                reg.fields = self._parse_fields(module, reg, field_rows)
            elif field_rows:
                raise CSRValidationError(
                    f"{raw_name}: {reg_type} entries must not define fields"
                )

            module.registers.append(reg)
            previous = reg
            size = reg.byte_size(module.word_bytes)
            next_offset = offset + size

        self._infer_region_sizes(module)
        self._validate_shadow_depths(module)

    def _parse_fields(
        self,
        module: ModuleModel,
        reg: RegisterModel,
        rows: list[dict[str, Any]],
    ) -> list[FieldModel]:
        fields: list[FieldModel] = []
        used_bits = 0
        names: set[str] = set()
        for row in rows:
            row_number = int(row.get("__row__", 0))
            name = validate_identifier(
                self._cell(row, "field"),
                f"{reg.raw_name} row {row_number} field",
            )
            if name in names:
                raise CSRValidationError(
                    f"{reg.raw_name}: duplicate field '{name}'"
                )
            names.add(name)
            msb = parse_int(self._cell(row, "msb"), f"{reg.raw_name}.{name} msb")
            lsb = parse_int(self._cell(row, "lsb"), f"{reg.raw_name}.{name} lsb")
            if lsb < 0 or msb < lsb or msb >= module.base_info.reg_bitwidth:
                raise CSRValidationError(
                    f"{reg.raw_name}.{name}: invalid bit range [{msb}:{lsb}] "
                    f"for width {module.base_info.reg_bitwidth}"
                )
            mask = ((1 << (msb - lsb + 1)) - 1) << lsb
            if used_bits & mask:
                raise CSRValidationError(
                    f"{reg.raw_name}.{name}: bit range [{msb}:{lsb}] overlaps "
                    "another field"
                )
            used_bits |= mask
            row_access = self._cell(row, "sw_access").upper()
            if row_access and row_access != reg.sw_access:
                raise CSRValidationError(
                    f"{reg.raw_name}.{name}: SW_access {row_access} differs "
                    f"from register access {reg.sw_access}"
                )
            width = msb - lsb + 1
            fields.append(FieldModel(
                name=name,
                msb=msb,
                lsb=lsb,
                sw_access=reg.sw_access,
                default_values=expand_defaults(
                    self._cell(row, "default_value"),
                    reg.repeat,
                    width,
                    f"{reg.raw_name}.{name} default_value",
                ),
                description=self._cell(row, "description"),
            ))
        return fields

    def _validate_special(
        self, name: str, reg_type: str, special: SpecialOptions
    ) -> None:
        if special.extras:
            raise CSRValidationError(
                f"{name}: unsupported special option(s): "
                + ", ".join(special.extras)
            )
        if special.shadow and reg_type != "cfg":
            raise CSRValidationError(
                f"{name}: shadow is only valid for reg_type=cfg"
            )
        if reg_type == "slave" and not special.slv_filename:
            raise CSRValidationError(
                f"{name}: reg_type=slave requires slv_filename"
            )
        if reg_type != "slave" and special.slv_filename:
            raise CSRValidationError(
                f"{name}: slv_filename is only valid for reg_type=slave"
            )
        if reg_type == "mem" and special.bytesize is None:
            raise CSRValidationError(
                f"{name}: reg_type=mem requires bytesize"
            )
        if reg_type not in {"slave", "mem"} and special.bytesize is not None:
            raise CSRValidationError(
                f"{name}: bytesize is only valid for slave or mem"
            )
        if reg_type in {"slave", "mem"} and special.repeat != 1:
            raise CSRValidationError(
                f"{name}: repeat is not supported for slave or mem"
            )

    def _infer_region_sizes(self, module: ModuleModel) -> None:
        for index, reg in enumerate(module.registers):
            if reg.reg_type not in {"slave", "mem"}:
                continue
            if reg.special.bytesize is None:
                if index + 1 < len(module.registers):
                    reg.special.bytesize = (
                        module.registers[index + 1].offset - reg.offset
                    )
                elif module.base_info.system_bytesize is not None:
                    reg.special.bytesize = (
                        module.base_info.system_bytesize - reg.offset
                    )
                else:
                    raise CSRValidationError(
                        f"{reg.raw_name}: bytesize cannot be inferred; set "
                        "bytesize or system_bytesize"
                    )
            if reg.special.bytesize <= 0:
                raise CSRValidationError(
                    f"{reg.raw_name}: inferred bytesize must be positive"
                )

    def _validate_shadow_depths(self, module: ModuleModel) -> None:
        depths = {
            reg.special.shadow
            for reg in module.registers
            if reg.special.shadow >= 2
        }
        if len(depths) > 1:
            raise CSRValidationError(
                f"{module.name}: all shadow N depths >= 2 must match"
            )

    def _load_children(self, module: ModuleModel, parent_path: Path) -> None:
        for reg in module.registers:
            if reg.reg_type != "slave":
                continue
            child_path = (parent_path.parent / reg.special.slv_filename).resolve()
            if not child_path.exists():
                raise FileNotFoundError(
                    f"{reg.raw_name}: slave file not found: {child_path}"
                )
            child = self._parse_file(child_path, reg.special.bytesize)
            module.sub_modules.append(SubModuleNode(
                instance_name=reg.name,
                offset=reg.offset,
                bytesize=reg.special.bytesize or 0,
                source_path=str(child_path),
                module_obj=child,
            ))

    @staticmethod
    def _cell(row: dict[str, Any], name: str) -> str:
        value = row.get(name, "")
        if value is None:
            return ""
        if isinstance(value, list):
            return ",".join(str(item) for item in value)
        return str(value).strip()
