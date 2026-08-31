from __future__ import annotations

from pathlib import Path


SUPPORTED_FORMATS = {"md", "xlsx"}


def create_template(
    output_path: str | Path,
    template_format: str | None = None,
    include_base_info: bool = False,
) -> Path:
    output = Path(output_path).resolve()
    selected_format = _resolve_format(output, template_format)
    output.parent.mkdir(parents=True, exist_ok=True)

    source = Path(__file__).resolve().parents[1] / "doc" / "reg_template.md"
    if selected_format == "md":
        _write_markdown_template(source, output, include_base_info)
    else:
        _write_excel_template(source, output, include_base_info)
    return output


def _resolve_format(output: Path, template_format: str | None) -> str:
    suffix_format = output.suffix.lower().lstrip(".")
    selected_format = (template_format or suffix_format or "md").lower()
    if selected_format not in SUPPORTED_FORMATS:
        raise ValueError(
            "Template format must be 'md' or 'xlsx'; "
            f"got '{selected_format}'"
        )
    if suffix_format and suffix_format in SUPPORTED_FORMATS:
        if template_format and suffix_format != selected_format:
            raise ValueError(
                f"Output suffix '.{suffix_format}' conflicts with "
                f"--format {selected_format}"
            )
    elif output.suffix:
        raise ValueError("Template output suffix must be '.md' or '.xlsx'")
    return selected_format


def _write_markdown_template(
    source: Path,
    output: Path,
    include_base_info: bool,
) -> None:
    tables = _read_markdown_tables(source)
    sections = []
    if include_base_info:
        sections.extend([
            "# base_info",
            "",
            _markdown_table(tables["base_info"]),
            "",
        ])
    sections.extend([
        "# reg_define",
        "",
        _markdown_table(tables["reg_define"]),
        "",
    ])
    output.write_text("\n".join(sections), encoding="utf-8")


def _write_excel_template(
    source: Path,
    output: Path,
    include_base_info: bool,
) -> None:
    try:
        import openpyxl
    except ImportError as exc:
        raise ImportError("openpyxl is required for .xlsx templates") from exc

    tables = _read_markdown_tables(source)
    workbook = openpyxl.Workbook()
    if include_base_info:
        base_sheet = workbook.active
        base_sheet.title = "base_info"
        _append_table(base_sheet, tables["base_info"])
        reg_sheet = workbook.create_sheet("reg_define")
    else:
        reg_sheet = workbook.active
        reg_sheet.title = "reg_define"
    _append_table(reg_sheet, tables["reg_define"])
    _add_list_validation(reg_sheet, "F2:F1048576", "RW,RO,W1T,W1C")
    _add_list_validation(reg_sheet, "H2:H1048576", "cfg,status,cmd,irq,slave,mem")

    _style_workbook(workbook)
    workbook.save(output)


def _read_markdown_tables(source: Path) -> dict[str, list[list[str]]]:
    sections: dict[str, list[str]] = {}
    current_section = ""
    for line in source.read_text(encoding="utf-8-sig").splitlines():
        stripped = line.strip()
        if stripped.startswith("#"):
            current_section = stripped.lstrip("#").strip().lower()
            sections.setdefault(current_section, [])
        elif current_section:
            sections[current_section].append(line)

    tables = {}
    for section in ("base_info", "reg_define"):
        rows = []
        for line in sections.get(section, []):
            stripped = line.strip()
            if not stripped.startswith("|"):
                continue
            cells = [cell.strip() for cell in stripped.strip("|").split("|")]
            if all(set(cell) <= {"-", ":", " "} for cell in cells):
                continue
            rows.append(cells)
        if not rows:
            raise ValueError(f"Template source is missing the '{section}' table")
        tables[section] = rows
    return tables


def _append_table(sheet: object, rows: list[list[str]]) -> None:
    for row in rows:
        sheet.append(row)


def _markdown_table(rows: list[list[str]]) -> str:
    widths = [
        max(len(row[index]) for row in rows)
        for index in range(len(rows[0]))
    ]

    def format_row(row: list[str]) -> str:
        return "| " + " | ".join(
            row[index].ljust(widths[index])
            for index in range(len(widths))
        ) + " |"

    separator = ["-" * width for width in widths]
    return "\n".join([
        format_row(rows[0]),
        format_row(separator),
        *(format_row(row) for row in rows[1:]),
    ])


def _add_list_validation(sheet: object, cell_range: str, values: str) -> None:
    from openpyxl.worksheet.datavalidation import DataValidation

    validation = DataValidation(
        type="list",
        formula1=f'"{values}"',
        sqref=cell_range,
    )
    validation.error = "Select a value from the list."
    validation.errorTitle = "Invalid value"
    validation.prompt = "Select a supported value."
    validation.promptTitle = "CSR template"
    validation.showErrorMessage = True
    validation.showInputMessage = True
    sheet.add_data_validation(validation)


def _style_workbook(workbook: object) -> None:
    from openpyxl.styles import Font, PatternFill
    from openpyxl.utils import get_column_letter

    for sheet in workbook.worksheets:
        for cell in sheet[1]:
            cell.font = Font(bold=True)
            cell.fill = PatternFill("solid", fgColor="DCE6F1")
        sheet.freeze_panes = "A2"
        sheet.auto_filter.ref = sheet.dimensions
        for column_index, column in enumerate(sheet.columns, start=1):
            width = min(
                60,
                max(12, max(len(str(cell.value or "")) for cell in column) + 2),
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = width
