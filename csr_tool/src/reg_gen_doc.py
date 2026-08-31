from __future__ import annotations

from collections import Counter, defaultdict
from pathlib import Path
from typing import Iterable

from .models import ModuleModel, RegisterModel
from .reg_common import write_text


REG_HEADERS = [
    "offset",
    "reg_name",
    "field",
    "msb",
    "lsb",
    "SW_access",
    "default_value",
    "reg_type",
    "special",
    "description",
]

TREE_REG_HEADERS = ["address", *REG_HEADERS[1:]]


class DocGenerator:
    def __init__(self, module: ModuleModel, out_dir: str):
        self.module = module
        self.out_dir = Path(out_dir)

    def generate_all(self, is_nested: bool = False) -> list[Path]:
        self.out_dir.mkdir(parents=True, exist_ok=True)
        generated: list[Path] = []
        modules = list(self.module.walk()) if is_nested else [
            (self.module, self.module.base_info.system_baseaddr, (self.module.name,))
        ]

        for module, _, _ in modules:
            path = self.out_dir / f"{module.name}_gen.md"
            write_text(path, self.module_markdown(module))
            generated.append(path)

        if is_nested:
            tree_md = self.out_dir / f"{self.module.name}_tree.md"
            tree_html = self.out_dir / f"{self.module.name}_tree.html"
            write_text(tree_md, self.tree_markdown())
            generated.append(tree_md)
            try:
                write_text(tree_html, self.tree_html())
            except ImportError:
                print("[Warning] jinja2 is unavailable; skipping HTML output")
            else:
                generated.append(tree_html)
            excel = self._write_tree_excel()
        else:
            excel = self._write_module_excel(self.module)
        if excel is not None:
            generated.append(excel)

        return generated

    def module_markdown(self, module: ModuleModel) -> str:
        base = module.base_info
        base_rows = [
            ["reg_bitwidth", str(base.reg_bitwidth), "-"],
            ["system_baseaddr", f"0x{base.system_baseaddr:X}", "-"],
        ]
        if base.system_bytesize is not None:
            base_rows.append(["system_bytesize", f"0x{base.system_bytesize:X}", "-"])
        if base.system_prefix:
            base_rows.append(["system_prefix", base.system_prefix, "-"])
        if base.author:
            base_rows.append(["author", base.author, "-"])
        if base.email:
            base_rows.append(["email", base.email, "-"])
        base_rows.extend([key, value, "-"] for key, value in base.extras.items())

        parts = [
            "# base_info",
            "",
            self._markdown_table(["item", "type_input", "description"], base_rows),
            "",
            "# reg_define",
            "",
            self._markdown_table(REG_HEADERS, self._register_rows(module)),
        ]
        return "\n".join(parts)

    def tree_markdown(self) -> str:
        nodes = self._document_nodes()
        lines = [
            f"# {self.module.name} Register Tree",
            "",
            "## Address Map",
            "",
            self._markdown_table(
                ["block", "address_range", "bytesize"],
                [
                    [
                        f"{section_number} {display_name}",
                        self._address_range(base, size),
                        f"0x{size:X}",
                    ]
                    for module, base, _, size, section_number, display_name in nodes
                ],
            ),
        ]
        for module, base, _, _, section_number, display_name in nodes:
            lines.extend([
                "",
                f"## {section_number} {display_name}",
                "",
                self._markdown_table(
                    TREE_REG_HEADERS,
                    self._tree_register_rows(module, base),
                ),
            ])
        return "\n".join(lines)

    def tree_html(self) -> str:
        template = self._jinja_env().get_template("tree.html.j2")
        return template.render(
            module_name=self.module.name,
            nodes=self._html_document_nodes(),
        )

    @staticmethod
    def _jinja_env():
        from jinja2 import Environment, FileSystemLoader, select_autoescape

        template_dir = Path(__file__).resolve().parent
        return Environment(
            loader=FileSystemLoader(template_dir),
            autoescape=select_autoescape(["html", "j2"]),
            trim_blocks=True,
            lstrip_blocks=True,
        )

    def _html_document_nodes(self) -> list[dict[str, object]]:
        nodes = self._document_nodes()
        html_nodes: list[dict[str, object]] = []
        for module, base, path, size, section_number, display_name in nodes:
            anchor = f"module-{section_number.replace('.', '-')}"
            registers = []
            for reg_index, reg in enumerate(module.registers, start=1):
                registers.append({
                    "index": reg_index,
                    "anchor": f"{anchor}-reg-{reg_index}",
                    "name": reg.name,
                    "address": f"0x{base + reg.offset:X}",
                    "reg_type": reg.reg_type,
                    "special": reg.special.to_text(),
                    "sw_access": reg.sw_access or "-",
                    "fields": [
                        {
                            "name": field.name,
                            "bit_scope": f"[{field.msb}:{field.lsb}]",
                            "default_value": field.default_value,
                            "description": field.description or reg.description,
                        }
                        for field in reg.fields
                    ],
                })
            html_nodes.append({
                "anchor": anchor,
                "depth": len(path) - 1,
                "section_number": section_number,
                "display_name": display_name,
                "address_range": self._address_range(base, size),
                "bytesize": f"0x{size:X}",
                "registers": registers,
            })
        return html_nodes

    def _register_rows(self, module: ModuleModel) -> list[list[str]]:
        rows: list[list[str]] = []
        for reg in module.registers:
            if not reg.fields:
                rows.append([
                    f"0x{reg.offset:X}",
                    reg.name,
                    "",
                    "",
                    "",
                    "",
                    "",
                    reg.reg_type,
                    reg.special.to_text(),
                    reg.description,
                ])
                continue
            for field_index, field in enumerate(reg.fields):
                first = field_index == 0
                rows.append([
                    f"0x{reg.offset:X}" if first else "",
                    reg.name if first else "",
                    field.name,
                    str(field.msb),
                    str(field.lsb),
                    reg.sw_access if first else "",
                    field.default_value,
                    reg.reg_type if first else "",
                    reg.special.to_text() if first else "",
                    field.description or (reg.description if first else ""),
                ])
        return rows

    def _tree_register_rows(
        self,
        module: ModuleModel,
        absolute_base: int,
    ) -> list[list[str]]:
        rows = self._register_rows(module)
        for row in rows:
            if row[0]:
                row[0] = f"0x{absolute_base + int(row[0], 0):X}"
        return rows

    def _document_nodes(
        self,
    ) -> list[tuple[ModuleModel, int, tuple[str, ...], int, str, str]]:
        raw_nodes: list[
            tuple[ModuleModel, int, tuple[str, ...], int, str]
        ] = []

        def visit(
            module: ModuleModel,
            absolute_base: int,
            path: tuple[str, ...],
            allocated_size: int,
            section_parts: tuple[int, ...],
        ) -> None:
            current_path = path + (module.name,)
            section_number = ".".join(str(item) for item in section_parts)
            raw_nodes.append(
                (
                    module,
                    absolute_base,
                    current_path,
                    allocated_size,
                    section_number,
                )
            )
            for child_index, child in enumerate(module.sub_modules, start=1):
                visit(
                    child.module_obj,
                    absolute_base + child.offset,
                    current_path,
                    child.bytesize,
                    section_parts + (child_index,),
                )

        root_size = self.module.base_info.system_bytesize or self.module.local_size
        visit(
            self.module,
            self.module.base_info.system_baseaddr,
            (),
            root_size,
            (1,),
        )
        name_counts = Counter(module.name for module, _, _, _, _ in raw_nodes)
        name_indexes: defaultdict[str, int] = defaultdict(int)
        nodes = []
        for module, base, path, size, section_number in raw_nodes:
            name_indexes[module.name] += 1
            display_name = module.name
            if name_counts[module.name] > 1:
                display_name = f"{module.name}_u{name_indexes[module.name]}"
            nodes.append(
                (
                    module,
                    base,
                    path,
                    size,
                    section_number,
                    display_name,
                )
            )
        return nodes

    @staticmethod
    def _address_range(start: int, size: int) -> str:
        end = start + max(1, size) - 1
        return f"0x{start:X} ~ 0x{end:X}"

    def _write_module_excel(self, module: ModuleModel) -> Path | None:
        try:
            import openpyxl
        except ImportError:
            print("[Warning] openpyxl is unavailable; skipping Excel output")
            return None
        path = self.out_dir / f"{module.name}_gen.xlsx"
        workbook = openpyxl.Workbook()
        base_sheet = workbook.active
        base_sheet.title = "base_info"
        base_sheet.append(["item", "type_input", "description"])
        base = module.base_info
        base_sheet.append(["reg_bitwidth", base.reg_bitwidth, "-"])
        base_sheet.append(["system_baseaddr", f"0x{base.system_baseaddr:X}", "-"])
        if base.system_bytesize is not None:
            base_sheet.append(["system_bytesize", f"0x{base.system_bytesize:X}", "-"])
        if base.system_prefix:
            base_sheet.append(["system_prefix", base.system_prefix, "-"])
        if base.author:
            base_sheet.append(["author", base.author, "-"])
        if base.email:
            base_sheet.append(["email", base.email, "-"])
        reg_sheet = workbook.create_sheet("reg_define")
        reg_sheet.append(REG_HEADERS)
        for row in self._register_rows(module):
            reg_sheet.append(row)
        self._style_workbook(workbook)
        workbook.save(path)
        return path

    def _write_tree_excel(self) -> Path | None:
        try:
            import openpyxl
        except ImportError:
            print("[Warning] openpyxl is unavailable; skipping Excel output")
            return None
        path = self.out_dir / f"{self.module.name}_tree.xlsx"
        workbook = openpyxl.Workbook()
        map_sheet = workbook.active
        map_sheet.title = "address_map"
        map_sheet.append(["block", "address_range", "bytesize", "link"])
        used_names = {"address_map"}
        for (
            module,
            base,
            _,
            size,
            section_number,
            display_name,
        ) in self._document_nodes():
            sheet_name = self._unique_sheet_name(
                f"{section_number.replace('.', '_')}_{display_name}",
                used_names,
            )
            sheet = workbook.create_sheet(sheet_name)
            sheet.append(TREE_REG_HEADERS)
            for row in self._tree_register_rows(module, base):
                sheet.append(row)
            map_sheet.append([
                f"{section_number} {display_name}",
                self._address_range(base, size),
                f"0x{size:X}",
                f"Open {sheet_name}",
            ])
            map_sheet.cell(row=map_sheet.max_row, column=4).hyperlink = f"#'{sheet_name}'!A1"
            map_sheet.cell(row=map_sheet.max_row, column=4).style = "Hyperlink"
        self._style_workbook(workbook)
        workbook.save(path)
        return path

    @staticmethod
    def _style_workbook(workbook: object) -> None:
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter

        for sheet in workbook.worksheets:
            for cell in sheet[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill("solid", fgColor="DCE6F1")
            sheet.freeze_panes = "A2"
            sheet.auto_filter.ref = sheet.dimensions
            for column_index, column in enumerate(sheet.columns, start=1):
                width = min(48, max(10, max(len(str(cell.value or "")) for cell in column) + 2))
                sheet.column_dimensions[get_column_letter(column_index)].width = width

    @staticmethod
    def _unique_sheet_name(name: str, used: set[str]) -> str:
        base = name[:31]
        candidate = base
        suffix = 1
        while candidate in used:
            tail = f"_{suffix}"
            candidate = base[: 31 - len(tail)] + tail
            suffix += 1
        used.add(candidate)
        return candidate

    @staticmethod
    def _markdown_table(headers: list[str], rows: Iterable[list[str]]) -> str:
        rows_list = [[str(item) for item in row] for row in rows]
        widths = [
            max(len(headers[index]), *(len(row[index]) for row in rows_list))
            for index in range(len(headers))
        ]
        def line(values: list[str]) -> str:
            return "| " + " | ".join(
                values[index].ljust(widths[index])
                for index in range(len(headers))
            ) + " |"
        return "\n".join([
            line(headers),
            "| " + " | ".join("-" * width for width in widths) + " |",
            *(line(row) for row in rows_list),
        ])
