from __future__ import annotations

import tempfile
import unittest
from pathlib import Path

from src.autogen_reg import run
from src.reg_common import CSRValidationError
from src.reg_parser import CSRParser


ROOT = Path(__file__).resolve().parent


class ParserTests(unittest.TestCase):
    def test_nested_address_map_and_name_deduplication(self) -> None:
        module = CSRParser(str(ROOT / "input" / "top_reg.md"), nested=True).parse()
        self.assertEqual(module.name, "top")
        self.assertEqual(
            [reg.name for reg in module.registers[-4:]],
            ["test1", "test2", "test3", "test4"],
        )
        addresses = [
            ("/".join(path), base) for _, base, path in module.walk()
        ]
        self.assertIn(("top/mid_a/leaf_a1", 0xF0001100), addresses)
        self.assertIn(("top/mid_b/leaf_a2", 0xF0002100), addresses)

    def test_generated_markdown_is_idempotent(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            generated = run(
                str(ROOT / "input" / "leaf_a1_reg.md"),
                temp,
                nested=False,
            )
            markdown = Path(temp) / "doc" / "leaf_a1_gen.md"
            original = CSRParser(str(ROOT / "input" / "leaf_a1_reg.md")).parse()
            reparsed = CSRParser(str(markdown)).parse()
            self.assertEqual(
                [
                    (reg.name, reg.offset, reg.reg_type, reg.default_word())
                    for reg in original.registers
                ],
                [
                    (reg.name, reg.offset, reg.reg_type, reg.default_word())
                    for reg in reparsed.registers
                ],
            )
            self.assertIn(Path(temp) / "rtl" / "leaf_a1.sv", generated)

    def test_overlap_is_rejected(self) -> None:
        text = """# reg_define

| offset | reg_name | field | msb | lsb | SW_access | default_value | reg_type | special | description |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| 0x0 | first | value | 31 | 0 | RW | 0 | cfg | repeat 2 | |
| 0x4 | second | value | 31 | 0 | RW | 0 | cfg | - | |
"""
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "bad.md"
            path.write_text(text, encoding="utf-8")
            with self.assertRaisesRegex(CSRValidationError, "Address overlap"):
                CSRParser(str(path)).parse()

    def test_reserved_field_name_is_rejected(self) -> None:
        text = """# reg_define

| offset | reg_name | field | msb | lsb | SW_access | default_value | reg_type | special | description |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| 0x0 | first | while | 0 | 0 | RW | 0 | cfg | - | |
"""
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "bad.md"
            path.write_text(text, encoding="utf-8")
            with self.assertRaisesRegex(CSRValidationError, "reserved keyword"):
                CSRParser(str(path)).parse()

    def test_vhdl_reserved_field_name_is_rejected(self) -> None:
        text = """# reg_define

| offset | reg_name | field | msb | lsb | SW_access | default_value | reg_type | special | description |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| 0x0 | first | architecture | 0 | 0 | RW | 0 | cfg | - | |
"""
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "bad.md"
            path.write_text(text, encoding="utf-8")
            with self.assertRaisesRegex(
                CSRValidationError,
                "'architecture' is a reserved keyword",
            ):
                CSRParser(str(path)).parse()

    def test_json_input_is_rejected(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            path = Path(temp) / "unsupported.json"
            path.write_text("{}", encoding="utf-8")
            with self.assertRaisesRegex(
                CSRValidationError,
                "unsupported input format '.json'",
            ):
                CSRParser(str(path)).parse()

    def test_shadow_and_irq_generation(self) -> None:
        text = """# base_info

| item | type_input |
| --- | --- |
| reg_bitwidth | 32 |

# reg_define

| offset | reg_name | field | msb | lsb | SW_access | default_value | reg_type | special | description |
| --- | --- | --- | --- | --- | --- | --- | --- | --- | --- |
| 0x0 | cfg_queue | value | 7 | 0 | RW | 1,2 | cfg | repeat 2, shadow 4 | |
| 0x8 | irq_state | done | 0 | 0 | W1C | 0 | irq | - | |
"""
        with tempfile.TemporaryDirectory() as temp:
            source = Path(temp) / "shadow_reg.md"
            source.write_text(text, encoding="utf-8")
            run(str(source), temp, nested=False)
            rtl = (Path(temp) / "rtl" / "shadow.sv").read_text(encoding="utf-8")
            self.assertIn("localparam integer SHADOW_DEPTH = 4;", rtl)
            self.assertRegex(
                rtl,
                r"REG_CFG_QUEUE_0_ADDR\s+= CSR_AW'\(32'h00000000\);",
            )
            self.assertRegex(
                rtl,
                r"REG_CFG_QUEUE_1_ADDR\s+= CSR_AW'\(32'h00000004\);",
            )
            self.assertRegex(
                rtl,
                r"REG_IRQ_STATE_ADDR\s+= CSR_AW'\(32'h00000008\);",
            )
            self.assertIn("r_cfg_queue_value[r_shadow_wr_idx]", rtl)
            self.assertIn("assign o_irqclr_irq_state_done", rtl)
            self.assertIn(
                "b_local_write_fire && "
                "(i_csr_req_addr == REG_IRQ_STATE_ADDR)",
                rtl,
            )
            self.assertIn(
                "i_csr_req_addr == REG_IRQ_STATE_ADDR",
                rtl,
            )
            self.assertNotIn("i_csr_req_addr == CSR_AW'", rtl)
            self.assertIn("if (!rst_n) begin", rtl)
            self.assertIn("else if (clear) begin", rtl)
            self.assertNotIn("if (!rst_n || clear)", rtl)
            self.assertNotIn("always_ff", rtl)

    def test_generated_rtl_port_columns_are_aligned(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            run(
                str(ROOT / "input" / "top_reg.md"),
                temp,
                nested=True,
            )
            rtl = (Path(temp) / "rtl" / "top.sv").read_text(encoding="utf-8")
            self.assertIn("parameter CSR_AW = 32,", rtl)
            self.assertIn("parameter CSR_DW = 32", rtl)
            self.assertNotIn("parameter integer", rtl)
            port_block = rtl.split("(", 2)[2].split(");", 1)[0]
            port_lines = [
                line for line in port_block.splitlines()
                if line.strip().startswith(("input ", "output "))
            ]
            delimiter_columns = {
                line.index("//,") if "//," in line else line.index(",")
                for line in port_lines
            }
            self.assertEqual(len(delimiter_columns), 1)

            signal_columns = set()
            for line in port_lines:
                declaration = (
                    line.split(" //", 1)[0]
                    .split("//,", 1)[0]
                    .rsplit(",", 1)[0]
                )
                signal = declaration.split()[-1]
                signal_columns.add(line.index(signal))
            self.assertEqual(len(signal_columns), 1)

            tx_position = port_block.index("o_tx_csr_req_write")
            register_positions = [
                port_block.index("o_cmd_top_ctrl_start"),
                port_block.index("i_sta_top_ver_date"),
                port_block.index("o_cfg_test1_data"),
            ]
            self.assertTrue(
                all(tx_position < position for position in register_positions)
            )
            self.assertLess(
                port_block.index("clear"),
                port_block.index("i_csr_req_write"),
            )
            self.assertIn("if (!rst_n) begin", rtl)
            self.assertIn("else if (clear) begin", rtl)
            self.assertNotIn("if (!rst_n || clear)", rtl)
            self.assertIn(
                "if (!rst_n)\n"
                "        r_top_ctrl_start <= 1'h0;\n"
                "    else if (clear)\n"
                "        r_top_ctrl_start <= 1'h0;",
                rtl,
            )
            self.assertNotIn(
                "if (!rst_n) begin\n"
                "        r_top_ctrl_start <= 1'h0;",
                rtl,
            )
            self.assertIn(
                "assign o_csr_rsp_rvalid = w_rsp_rvalid && !clear;",
                rtl,
            )
            self.assertIn(
                "SLV_MID_A_ADDR_S = CSR_AW'(32'h00001000);",
                rtl,
            )
            self.assertIn(
                "SLV_MID_A_ADDR_E = CSR_AW'(32'h000017FF);",
                rtl,
            )
            self.assertIn(
                "i_csr_req_addr >= SLV_MID_A_ADDR_S",
                rtl,
            )
            self.assertIn(
                "i_csr_req_addr <= SLV_MID_A_ADDR_E",
                rtl,
            )
            self.assertIn(
                "if ((i_csr_req_addr >= SLV_MID_A_ADDR_S) && "
                "(i_csr_req_addr <= SLV_MID_A_ADDR_E))",
                rtl,
            )
            self.assertNotIn(
                "SLV_MID_A_ADDR_S) &&\n",
                rtl,
            )
            self.assertIn(
                "i_csr_req_addr - SLV_MID_A_ADDR_S",
                rtl,
            )
            self.assertIn("localparam integer SLV_SEL_W = 2;", rtl)
            self.assertIn(
                "localparam [SLV_SEL_W-1:0] SLV_LOCAL = 'd0;\n"
                "localparam [SLV_SEL_W-1:0] SLV_MID_A = 'd1;\n"
                "localparam [SLV_SEL_W-1:0] SLV_MID_B = 'd2;\n"
                "localparam [CSR_AW-1:0] SLV_MID_A_ADDR_S",
                rtl,
            )
            self.assertIn("w_req_slv", rtl)
            self.assertIn("r_read_slv", rtl)
            self.assertIn("b_slv_switch_block", rtl)
            self.assertNotIn("TARGET_", rtl)
            self.assertNotIn("w_req_target", rtl)
            self.assertIn(
                "assign o_tx_csr_req_valid[1] = i_csr_req_valid && "
                "!clear && !b_slv_switch_block && "
                "(w_req_slv == SLV_MID_B);",
                rtl,
            )
            self.assertNotIn(
                "o_tx_csr_req_valid[1] = i_csr_req_valid &&\n",
                rtl,
            )
            self.assertIn(
                "r_otf_cnt <= r_otf_cnt + "
                "b_read_fire - b_rsp_fire;",
                rtl,
            )
            self.assertIn(
                "else if (b_read_fire && (r_otf_cnt == '0))\n"
                "        r_read_slv <= w_req_slv;\n"
                "end\n\n"
                "always @(posedge clk or negedge rst_n) begin",
                rtl,
            )
            self.assertIn(
                "else if (b_read_fire || b_rsp_fire)\n"
                "        r_otf_cnt <= r_otf_cnt + "
                "b_read_fire - b_rsp_fire;",
                rtl,
            )
            self.assertNotIn(
                "r_read_slv <= w_req_slv;\n"
                "        r_otf_cnt <=",
                rtl,
            )
            self.assertIn(
                "CSR_INVALID_RDATA = CSR_DW'(32'hDEAFDEAF);",
                rtl,
            )
            self.assertIn(
                "w_local_rdata = CSR_INVALID_RDATA;",
                rtl,
            )
            self.assertIn(
                "w_rsp_rdata = CSR_INVALID_RDATA;",
                rtl,
            )
            self.assertIn(
                "always @* begin\n"
                "    case (r_read_slv)",
                rtl,
            )
            self.assertNotIn(
                "always @* begin\n"
                "    w_rsp_rdata = r_local_rsp_rdata;\n"
                "    w_rsp_rvalid = r_local_rsp_rvalid;\n"
                "    case (r_read_slv)",
                rtl,
            )
            self.assertIn(
                "always @* begin\n"
                "    w_local_rdata = '0;\n"
                "    case (i_csr_req_addr)\n"
                "        REG_TOP_CTRL_ADDR: begin",
                rtl,
            )
            self.assertNotIn(
                "if (i_csr_req_addr == REG_TOP_CTRL_ADDR)",
                rtl,
            )
            self.assertIn(
                "REG_TOP_VER_ADDR: begin",
                rtl,
            )
            self.assertIn(
                "default: w_local_rdata = CSR_INVALID_RDATA;",
                rtl,
            )
            self.assertLess(
                rtl.index("REG_TOP_CTRL_ADDR: begin"),
                rtl.index("w_local_rdata = CSR_INVALID_RDATA;"),
            )
            self.assertIn(
                "for (int byte_idx = 0; "
                "byte_idx < (CSR_DW / 8); "
                "byte_idx = byte_idx + 1) begin",
                rtl,
            )
            self.assertNotIn("integer byte_idx;", rtl)
            self.assertNotIn(
                "case ({b_read_fire, b_rsp_fire})",
                rtl,
            )
            self.assertIn(
                "o_tx_csr_req_write",
                rtl,
            )
            self.assertIn("// [0]=mid_a, [1]=mid_b", rtl)
            self.assertRegex(
                rtl,
                r"REG_TOP_CTRL_ADDR\s+= CSR_AW'\(32'h00000000\);",
            )
            self.assertIn(
                "CSR_INVALID_RDATA = CSR_DW'(32'hDEAFDEAF);\n\n"
                "localparam [CSR_AW-1:0] REG_TOP_CTRL_ADDR",
                rtl,
            )
            self.assertRegex(
                rtl,
                r"REG_TEST1_0_ADDR\s+= CSR_AW'\(32'h00003000\);",
            )
            self.assertRegex(
                rtl,
                r"REG_TEST1_3_ADDR\s+= CSR_AW'\(32'h0000300C\);",
            )
            reg_addr_lines = [
                line for line in rtl.splitlines()
                if line.startswith("localparam [CSR_AW-1:0] REG_")
            ]
            self.assertEqual(
                len({line.index("=") for line in reg_addr_lines}),
                1,
            )
            self.assertIn(
                "i_csr_req_addr == REG_TOP_CTRL_ADDR",
                rtl,
            )
            self.assertIn(
                "if (b_local_write_fire && "
                "(i_csr_req_addr == REG_TOP_CTRL_ADDR)) begin",
                rtl,
            )
            self.assertIn(
                "r_test1_data[0] <= i_csr_req_wdata[31:0] & "
                "w_csr_wmask[31:0];",
                rtl,
            )
            self.assertNotIn(
                "r_test1_data[0] & ~w_csr_wmask",
                rtl,
            )
            self.assertNotIn(
                "r_test1_data[0] <= i_csr_req_wdata[31:0];",
                rtl,
            )
            self.assertIn(
                "assign b_local_write_fire = "
                "b_req_fire && i_csr_req_write &&",
                rtl,
            )
            self.assertNotIn(
                "if (b_req_fire && i_csr_req_write",
                rtl,
            )
            self.assertIn(
                "assign b_req_fire = "
                "i_csr_req_valid && o_csr_req_ready;",
                rtl,
            )
            self.assertIn(
                "assign b_local_read_fire = b_read_fire &&",
                rtl,
            )
            self.assertNotIn("_accept", rtl)
            self.assertNotIn("i_csr_req_addr == CSR_AW'", rtl)

            mid_b_rtl = (
                Path(temp) / "rtl" / "mid_b.sv"
            ).read_text(encoding="utf-8")
            self.assertIn("SLV_LEAF_A2_ADDR_S", mid_b_rtl)
            self.assertIn("SLV_LEAF_A2_ADDR_E", mid_b_rtl)
            self.assertNotIn("SLV_B_SLV", mid_b_rtl)
            self.assertIn("// [0]=b_mem, [1]=leaf_a2", mid_b_rtl)

            wrapper = (
                Path(temp) / "rtl" / "plus" / "top_wrap.sv"
            ).read_text(encoding="utf-8")
            self.assertLess(
                wrapper.index("o_tx_csr_req_write"),
                wrapper.index("o_cfg"),
            )
            self.assertIn("input  wire", wrapper)
            self.assertIn("clear", wrapper)
            self.assertNotIn("parameter integer", wrapper)
            self.assertIn("// [0]=mid_a, [1]=mid_b", wrapper)

    def test_tree_html_uses_register_detail_tables(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            run(
                str(ROOT / "input" / "top_reg.md"),
                temp,
                nested=True,
            )
            page = (Path(temp) / "doc" / "top_tree.html").read_text(
                encoding="utf-8"
            )
            self.assertIn('class="register-table"', page)
            for heading in (
                "reg_name",
                "address",
                "reg_type",
                "special",
                "SW_access",
                "field",
                "bit_scope",
                "default_value",
                "description",
            ):
                self.assertIn(f">{heading}<", page)
            self.assertNotIn(">SW<", page)
            self.assertNotIn(">HW<", page)
            self.assertNotIn(">msb<", page)
            self.assertNotIn(">lsb<", page)
            self.assertIn(">[31:0]<", page)
            self.assertEqual(page.count('<col style="width:20ch">'), 3 * 25)
            self.assertEqual(page.count('<col style="width:60ch">'), 25)
            self.assertIn('<aside class="sidebar">', page)
            self.assertIn('class="page-shell sidebar-collapsed"', page)
            self.assertIn('data-testid="sidebar-toggle"', page)
            self.assertIn('<a class="nav-address" href="#address-map">', page)
            self.assertIn('<details class="nav-module-group"', page)
            self.assertIn(
                '<summary class="nav-module">1 top</summary>',
                page,
            )
            self.assertIn(
                '<summary class="nav-module">1.1 mid_a</summary>',
                page,
            )
            self.assertIn(
                '<summary class="nav-module">1.1.1 leaf_a1</summary>',
                page,
            )
            self.assertIn(
                '<summary class="nav-module">1.1.2 leaf_a2_u1</summary>',
                page,
            )
            self.assertIn(
                '<summary class="nav-module">1.2.1 leaf_a2_u2</summary>',
                page,
            )
            self.assertNotIn("1.1 top/mid_a", page)
            self.assertNotIn("nav-overview", page)
            self.assertNotIn(">Overview<", page)
            self.assertIn('href="#module-1-reg-1"', page)
            self.assertIn('id="module-1-reg-1"', page)
            self.assertIn("top_ctrl (0xF0000000)", page)
            self.assertIn(
                '<h3 class="register-title" id="module-1-reg-1">'
                "top_ctrl (0xF0000000)</h3>",
                page,
            )
            self.assertIn(
                '<th class="meta-label">reg_name</th>'
                '<td class="meta-value">top_ctrl</td>',
                page,
            )
            self.assertNotIn(
                '<td class="meta-value">top_ctrl (0xF0000000)</td>',
                page,
            )
            self.assertIn("0xF0000000 ~ 0xF000FFFF", page)
            self.assertIn("0xF0001000 ~ 0xF00017FF", page)
            self.assertIn(">bytesize<", page)
            self.assertIn(">0x800<", page)
            self.assertIn(">block<", page)
            self.assertIn(">1.1 mid_a<", page)
            self.assertNotIn(">path<", page)
            self.assertNotIn("top/mid_a", page)
            self.assertNotIn(">source<", page)
            self.assertNotIn(">base_address<", page)
            self.assertIn(">address<", page)
            self.assertIn(">0xF0001000<", page)
            self.assertNotIn(">offset<", page)
            self.assertIn("<summary>1.1 mid_a</summary>", page)
            self.assertNotIn("<code>0xF0001000</code>", page)
            self.assertIn('class="address-table"', page)
            self.assertEqual(page.count('<col style="width:30ch">'), 3)

    def test_tree_markdown_uses_absolute_addresses(self) -> None:
        with tempfile.TemporaryDirectory() as temp:
            run(
                str(ROOT / "input" / "top_reg.md"),
                temp,
                nested=True,
            )
            tree = (Path(temp) / "doc" / "top_tree.md").read_text(
                encoding="utf-8"
            )
            self.assertIn("| block", tree)
            self.assertIn("| 1 top", tree)
            self.assertIn("| 1.1 mid_a", tree)
            self.assertIn("| 1.1.2 leaf_a2_u1", tree)
            self.assertIn("| 1.2.1 leaf_a2_u2", tree)
            self.assertNotIn("top/mid_a", tree)
            self.assertNotIn("Base address:", tree)
            self.assertNotIn("Source:", tree)
            self.assertIn("| address", tree)
            self.assertIn("| 0xF0000000 | top_ctrl", tree)
            self.assertIn("| 0xF0001000 | mid_cfg", tree)
            self.assertIn("| 0xF0001100 | ver", tree)

    def test_tree_excel_uses_absolute_addresses(self) -> None:
        try:
            import openpyxl
        except ImportError:
            self.skipTest("openpyxl is unavailable")
        with tempfile.TemporaryDirectory() as temp:
            run(
                str(ROOT / "input" / "top_reg.md"),
                temp,
                nested=True,
            )
            workbook = openpyxl.load_workbook(
                Path(temp) / "doc" / "top_tree.xlsx",
                read_only=True,
                data_only=True,
            )
            try:
                address_rows = list(workbook["address_map"].values)
                self.assertEqual(
                    address_rows[0],
                    ("block", "address_range", "bytesize", "link"),
                )
                self.assertEqual(address_rows[2][0], "1.1 mid_a")
                self.assertNotIn("top/mid_a", address_rows[2])
                self.assertEqual(address_rows[4][0], "1.1.2 leaf_a2_u1")
                self.assertEqual(address_rows[6][0], "1.2.1 leaf_a2_u2")
                self.assertIn("1_1_2_leaf_a2_u1", workbook.sheetnames)
                self.assertIn("1_2_1_leaf_a2_u2", workbook.sheetnames)

                mid_rows = list(workbook["1_1_mid_a"].values)
                self.assertEqual(mid_rows[0][0], "address")
                self.assertEqual(mid_rows[1][0], "0xF0001000")
                self.assertEqual(mid_rows[1][1], "mid_cfg")
            finally:
                workbook.close()


if __name__ == "__main__":
    unittest.main()
